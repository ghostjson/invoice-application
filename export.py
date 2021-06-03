from openpyxl import Workbook, load_workbook
from database import *
from shutil import copyfile
import json
from openpyxl.styles.borders import Border, Side

INVOICE_TEMPLATE = 'resources/templates/invoice_template.xlsx'

# =================================INVOICE SHEET================================ #

INVOICE_CLIENT = 'B1'
INVOICE_NAME = 'B2'
INVOICE_USER = 'B3'
INVOICE_SMV = 'B4'
INVOICE_LOCATION = 'B5'

# DEMO INVOICE ID
INVOICE_ID = 5


INVOICE_ITEM_ID = 'A8'
INVOICE_ITEM_NAME = 'B8'
INVOICE_ITEM_QUANTITY = 'C8'
INVOICE_ITEM_UNIT = 'D8'
INVOICE_ITEM_UNIT_VALUE = 'E8'
INVOICE_ITEM_TOTAL_VALUE = 'F8'

# =============================================================================== #

# =================================ITEM DETAIL PAGE============================== #
ITEM_NAME = 'A4'
ITEM_PERFORMANCE = 'E5'
ITEM_UNIT = 'G5'

# =============================================================================== #
BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
# =============================================================================== #


# apply border
def applyBorder(sheet, row, col_start, col_end):

	for col in range(col_start, col_end+1):
		sheet.cell(row, col).border = BORDER

# return total price of an item
def getItemPrice(item_id):

	sqls = ['','', '']

	sqls[0] = "SELECT SUM(price*quantity) AS total_equipment_price FROM Equipments \
		 			INNER JOIN ItemsEquipments \
					ON Equipments.equipment_id = ItemsEquipments.equipment_id \
					WHERE ItemsEquipments.item_id={}".format(item_id)
	sqls[1] = "SELECT SUM(price*quantity) AS total_material_price FROM Materials \
		 			INNER JOIN ItemsMaterials \
					ON Materials.material_id = ItemsMaterials.material_id \
					WHERE ItemsMaterials.item_id={}".format(item_id)
	sqls[2] = "SELECT SUM(hourWage*quantity) AS total_workforce_price FROM WorkForces \
					INNER JOIN ItemsWorkForces \
					ON WorkForces.workForce_id = ItemsWorkForces.workForce_id \
					WHERE ItemsWorkForces.item_id={}".format(item_id)

	total = 0
	for sql in sqls:
		price = list(json.loads(dbGET(sql))[0].values())[0]
		if price != None:
			total += price

	return total

# set invoice sheet
def setInvoiceSheet(invoice_data, save_location):


	invoice = load_workbook(save_location)
	invoice_sheet = invoice['invoice']

	invoice_sheet[INVOICE_NAME] = invoice_data['name']
	invoice_sheet[INVOICE_CLIENT] = invoice_data['client']
	invoice_sheet[INVOICE_USER] = invoice_data['user']
	invoice_sheet[INVOICE_SMV] = invoice_data['SMV']
	invoice_sheet[INVOICE_LOCATION] = invoice_data['location']



	invoice_items = invoiceItems(invoice_data['invoice_id'])
	invoice_items_len = len(invoice_items)

	# Entering each item details to the sheet
	for invoice_item in invoice_items:
		invoice_sheet.insert_rows(8)

		invoice_sheet[INVOICE_ITEM_ID] = invoice_item['invoice_id']
		invoice_sheet[INVOICE_ITEM_NAME] = invoice_item['name']
		invoice_sheet[INVOICE_ITEM_QUANTITY] = invoice_item['quantity']
		invoice_sheet[INVOICE_ITEM_UNIT] = invoice_item['unit']
		invoice_sheet[INVOICE_ITEM_UNIT_VALUE] = getItemPrice(invoice_item['item_id'])
		invoice_sheet[INVOICE_ITEM_TOTAL_VALUE] = getItemPrice(invoice_item['item_id']) * float(invoice_item['quantity'])


	INVOICE_SUB_TOTAL = 'F' + str(8+invoice_items_len)
	INVOICE_TAX = 'F' + str(9+invoice_items_len)
	INVOICE_PROFIT = 'F' + str(10+invoice_items_len)
	INVOICE_TOTAL = 'F' + str(11+invoice_items_len)

	# calculate subtotal
	invoice_sheet[INVOICE_SUB_TOTAL] = "=SUM(F8:F{})".format((7+invoice_items_len))
	invoice_sheet[INVOICE_TAX] = "={}*E{}".format(INVOICE_SUB_TOTAL, str(9+invoice_items_len))
	invoice_sheet[INVOICE_PROFIT] = "={}*E{}".format(INVOICE_TAX, str(10+invoice_items_len))
	invoice_sheet[INVOICE_TOTAL] = "=SUM({}:{})".format(INVOICE_SUB_TOTAL, INVOICE_PROFIT)


	invoice.save(save_location)

def setItemsDetailSheet(invoice_data, save_location):
	invoice = load_workbook(save_location)
	detail_template = invoice['item_detail']

	items = invoiceItems(invoice_data['invoice_id'])

	for item in items:

		# create sheet for each item
		invoice.copy_worksheet(detail_template)
		current_sheet = invoice['item_detail Copy']
		current_sheet.title = str(item['item_id'])

		# insert equipments to the item sheet
		get_equipment_query = "SELECT * FROM Equipments INNER JOIN ItemsEquipments \
													 ON Equipments.equipment_id = ItemsEquipments.equipment_id	\
													 WHERE ItemsEquipments.item_id={}".format(item['item_id'])
		equipments = json.loads(dbGET(get_equipment_query))

		no_of_equipments = len(equipments)

		if no_of_equipments > 0:
			for equipment in equipments:
				current_sheet.insert_rows(9)

				current_sheet['A9'] = equipment['name']
				# current_sheet['A9'].border = BORDER
				current_sheet['C9'] = equipment['quantity']
				current_sheet['D9'] = equipment['price']
				current_sheet['E9'] = equipment['price']
				current_sheet['G9'] = equipment['price'] * equipment['quantity']

				applyBorder(current_sheet, 9, 1, 7)

			current_sheet['G{}'.format(9+no_of_equipments)] = "=SUM(G9:G{})".format(8+no_of_equipments)
			total_equipment_price = current_sheet['G{}'.format(9+no_of_equipments)].value

		# insert workforce to the item sheet
		get_workforces_query = "SELECT * FROM Workforces INNER JOIN ItemsWorkforces \
														ON Workforces.workForce_id = ItemsWorkforces.workForce_id \
														WHERE ItemsWorkforces.item_id={}".format(item['item_id'])
		workforces = json.loads(dbGET(get_workforces_query))

		no_of_workforces = len(workforces)

		starting_row = 12 + no_of_equipments

		if no_of_workforces > 0:
			for workforce in workforces:
				current_sheet.insert_rows(starting_row)

				s = str(starting_row)

				current_sheet['A{}'.format(s)] = workforce['name']
				current_sheet['C{}'.format(s)] = workforce['quantity']
				current_sheet['D{}'.format(s)] = workforce['hourWage']
				current_sheet['E{}'.format(s)] = workforce['hourWage']
				current_sheet['G{}'.format(s)] = workforce['hourWage'] * workforce['quantity']


				applyBorder(current_sheet, starting_row, 1, 7)


			current_sheet['G'+ str(12+no_of_equipments+no_of_workforces)] = "=SUM(G{}:G{})".format(starting_row,starting_row+no_of_workforces - 1)
			total_workforce_price = current_sheet['G'+str(12+no_of_equipments+no_of_workforces)].value

		# insert materials to the item sheet
		get_material_query = "SELECT * FROM Materials INNER JOIN ItemsMaterials \
														ON Materials.material_id = ItemsMaterials.material_id \
														WHERE ItemsMaterials.item_id={}".format(item['item_id'])
		materials = json.loads(dbGET(get_material_query))

		no_of_materials = len(materials)

		starting_row = 15 + no_of_equipments + no_of_workforces

		if(no_of_materials > 0):
			for material in materials:
				current_sheet.insert_rows(starting_row)

				s = str(starting_row)

				current_sheet['A{}'.format(s)] = material['name']
				current_sheet['C{}'.format(s)] = material['quantity']
				current_sheet['D{}'.format(s)] = material['price']
				current_sheet['E{}'.format(s)] = material['price']
				current_sheet['G{}'.format(s)] = material['price'] * material['quantity']


				applyBorder(current_sheet, starting_row, 1, 7)


			current_sheet['G'+ str(starting_row+no_of_materials)] = "=SUM(G{}:G{})".format(starting_row,starting_row + no_of_materials - 1)
			total_material_price = current_sheet['G'+str(starting_row+no_of_materials)].value

		if no_of_materials == 0:
			current_sheet['G'+str(starting_row+no_of_materials+2)] = "=({}+{})".format(total_equipment_price[1:], total_workforce_price[1:])
		
		if no_of_equipments == 0:
			current_sheet['G'+str(starting_row+no_of_materials+2)] = "=({}+{})".format(total_material_price[1:], total_workforce_price[1:])
		
		if no_of_workforces == 0:
			current_sheet['G'+str(starting_row+no_of_materials+2)] = "=({}+{})".format(total_material_price[1:], total_equipment_price[1:])

		
		if no_of_materials != 0 and no_of_equipments != 0 and no_of_workforces !=0:
			current_sheet['G'+str(starting_row+no_of_materials+2)] = "=({}+{}+{})".format(total_material_price[1:], total_equipment_price[1:], total_workforce_price[1:])
		
		current_sheet['G'+str(starting_row+no_of_materials+3)] = "=({}*({})/100)".format(current_sheet['G'+str(starting_row+no_of_materials+2)].value[1:], 'F'+str(starting_row+no_of_materials+3))
		current_sheet['G'+str(starting_row+no_of_materials+7)].value = current_sheet['G'+str(starting_row+no_of_materials+8)].value  = "=SUM({}, {})".format(current_sheet['G'+str(starting_row+no_of_materials+3)].value[1:], current_sheet['G'+str(starting_row+no_of_materials+2)].value[1:])
		

	invoice.remove(detail_template)
	invoice.save(save_location)



# return items in a invoice
def invoiceItems(invoice_id):
	query = "SELECT * FROM Items INNER JOIN invoiceItems ON Items.item_id = invoiceItems.item_id WHERE invoiceItems.invoice_id='"+ str(invoice_id) +"'"

	invoice_items = json.loads(dbGET(query))
	return invoice_items

# export function
@eel.expose
def export(invoice_id, save_location):

	copyfile(INVOICE_TEMPLATE, save_location)

	invoice = load_workbook(save_location)

	invoice_sheet = invoice['invoice']
	invoices = dbGET('SELECT * FROM Invoices WHERE invoice_id='+ str(invoice_id))
	invoice_data = json.loads(invoices)[0]

	setInvoiceSheet(invoice_data, save_location)
	setItemsDetailSheet(invoice_data, save_location)

	return save_location


if __name__ == '__main__':
	export(INVOICE_ID, 'resources/exports/test.xlsx')
